import sys


sys.path.append("..")
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color,Border,Side
from tempfile import NamedTemporaryFile
from openpyxl.writer.excel import save_virtual_workbook 


def CreateNewFilebyOpertions(ilist,sheetName):
    stream=None
    wb = Workbook()
    #wsheet = wb.create_sheet(sheetName)
    wsheet = wb.active
    #wsheet = wb[sheetName]
    wsheet['A1'] = sheetName
    for x in ilist:
        wsheet[x] = ilist[x]

    # with NamedTemporaryFile() as tmp:
    #     wb.save(tmp.name)
    #     tmp.seek(0)
    #     stream = tmp.read()
    cont = save_virtual_workbook(wb)
    return cont


def CreateNewFile(ilist,sheetInfo):
    stream=None
    [sheetName,sheetTitle] = sheetInfo
    wb = Workbook()
    wsheet = wb.create_sheet(sheetName)
    wsheet['A1'] = sheetTitle
    for line in range(3 , len(ilist)+1):
        x=1
        wsheet.cell(x+1,line).value = line[x]
        x=x+1
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
         
    return stream
